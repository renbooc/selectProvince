from openpyxl import load_workbook

# 读取Excel文件
file_path = "7357e693383bb48ace55f9bd96fe6d8f.XLSX"

wb = load_workbook(file_path)
ws = wb.active

print("读取Excel文件内容：")
print("=" * 80)

# 读取所有数据
for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    # 清理数据
    cleaned_row = []
    for cell in row:
        if cell is None:
            cleaned_row.append("")
        else:
            cleaned_row.append(str(cell))

    if any(cleaned_row):  # 如果行不为空
        print(f"行{row_idx:2d}: {cleaned_row}")

print("\n" + "=" * 80)
print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")
